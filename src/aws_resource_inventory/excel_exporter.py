"""Excel Exporter for AWS Resource Inventory"""

import re
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from rich.console import Console

from aws_resource_inventory.scanners.base import ScanResult

console = Console()


def to_pascal_case(text: str) -> str:
    """Convert text to PascalCase for file naming"""
    # Remove special characters and split on spaces/underscores/hyphens
    words = re.split(r'[\s_\-]+', text)
    # Capitalize first letter of each word
    return ''.join(word.capitalize() for word in words if word)


class ExcelExporter:
    """Export scan results to individual Excel files per resource type"""
    
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(exist_ok=True)
    
    def export(
        self, 
        results: List[ScanResult], 
        errors: List[str],
        accounts_scanned: int,
        regions_scanned: int
    ) -> List[Path]:
        """
        Export all scan results to individual Excel files.
        
        Args:
            results: List of ScanResult objects from all scanners
            errors: List of error messages from scanning
            accounts_scanned: Number of accounts that were scanned
            regions_scanned: Number of regions per account
            
        Returns:
            List of paths to created Excel files
        """
        console.print(f"\n[yellow]📊 Exporting to Excel files...[/yellow]")
        
        created_files = []
        
        # Create individual files for each resource type with data
        for result in results:
            if result.resources:  # Only create file if there are resources
                file_path = self._create_resource_file(result, accounts_scanned, regions_scanned)
                created_files.append(file_path)
        
        # Create summary file
        summary_file = self._create_summary_file(results, errors, accounts_scanned, regions_scanned)
        created_files.append(summary_file)
        
        console.print(f"[green]✅ Created {len(created_files)} Excel files in {self.output_dir}[/green]")
        return created_files
    
    def _create_resource_file(
        self, 
        result: ScanResult, 
        accounts_scanned: int, 
        regions_scanned: int
    ) -> Path:
        """Create an individual Excel file for a resource type"""
        # Generate PascalCase filename
        filename = f"{to_pascal_case(result.sheet_name)}.xlsx"
        output_file = self.output_dir / filename
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df = pd.DataFrame(result.resources)
            df.to_excel(writer, index=False, sheet_name=result.sheet_name[:31])
            
            # Format the sheet
            worksheet = writer.sheets[result.sheet_name[:31]]
            self._format_resource_sheet(worksheet, df)
        
        console.print(f"  📁 {filename}: {len(result.resources)} resources")
        return output_file
    
    def _create_summary_file(
        self, 
        results: List[ScanResult], 
        errors: List[str],
        accounts_scanned: int,
        regions_scanned: int
    ) -> Path:
        """Create a summary Excel file with all resource counts"""
        output_file = self.output_dir / 'Summary.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Create Summary sheet
            self._create_summary_sheet(writer, results, accounts_scanned, regions_scanned)
            
            # Create Errors sheet
            self._create_errors_sheet(writer, errors)
        
        console.print(f"  📁 Summary.xlsx: Overview and {len(errors)} errors/skipped")
        return output_file
    
    def _create_summary_sheet(
        self, 
        writer: pd.ExcelWriter, 
        results: List[ScanResult],
        accounts_scanned: int,
        regions_scanned: int
    ):
        """Create the summary sheet with resource counts"""
        
        # Prepare summary data
        summary_data = []
        total_resources = 0
        
        for result in results:
            count = len(result.resources)
            total_resources += count
            summary_data.append({
                'Resource Type': result.resource_type,
                'Sheet Name': result.sheet_name,
                'Count': count,
                'Global': 'Yes' if result.is_global else 'No'
            })
        
        # Sort by count descending
        summary_data.sort(key=lambda x: x['Count'], reverse=True)
        
        # Add metadata rows at top
        metadata = [
            {'Resource Type': 'SCAN SUMMARY', 'Sheet Name': '', 'Count': '', 'Global': ''},
            {'Resource Type': 'Generated At', 'Sheet Name': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Count': '', 'Global': ''},
            {'Resource Type': 'Accounts Scanned', 'Sheet Name': str(accounts_scanned), 'Count': '', 'Global': ''},
            {'Resource Type': 'Regions Scanned', 'Sheet Name': str(regions_scanned), 'Count': '', 'Global': ''},
            {'Resource Type': 'Total Resources', 'Sheet Name': str(total_resources), 'Count': '', 'Global': ''},
            {'Resource Type': '', 'Sheet Name': '', 'Count': '', 'Global': ''},
            {'Resource Type': 'RESOURCE COUNTS', 'Sheet Name': '', 'Count': '', 'Global': ''},
        ]
        
        all_data = metadata + summary_data
        
        df = pd.DataFrame(all_data)
        df.to_excel(writer, index=False, sheet_name='Summary')
        
        # Format the sheet
        worksheet = writer.sheets['Summary']
        self._format_summary_sheet(worksheet, len(metadata))
    
    def _format_summary_sheet(self, worksheet, header_rows: int):
        """Apply formatting to summary sheet"""
        # Column widths
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 10
        
        # Header styling
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font_white = Font(bold=True, size=12, color='FFFFFF')
        
        # Style metadata section
        for row in range(1, header_rows + 1):
            cell = worksheet.cell(row=row, column=1)
            if cell.value in ['SCAN SUMMARY', 'RESOURCE COUNTS']:
                cell.font = Font(bold=True, size=14)
        
        # Style the actual header row (after metadata)
        data_header_row = header_rows + 1
        for col in range(1, 5):
            cell = worksheet.cell(row=data_header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font_white
            cell.alignment = Alignment(horizontal='center')
    
    def _create_resource_sheet(self, writer: pd.ExcelWriter, result: ScanResult):
        """Create a sheet for a specific resource type"""
        df = pd.DataFrame(result.resources)
        
        # Truncate sheet name to 31 characters (Excel limit)
        sheet_name = result.sheet_name[:31]
        
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        # Format the sheet
        worksheet = writer.sheets[sheet_name]
        self._format_resource_sheet(worksheet, df)
    
    def _format_resource_sheet(self, worksheet, df: pd.DataFrame):
        """Apply formatting to a resource sheet"""
        # Header styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Style header row
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for idx, col in enumerate(df.columns, 1):
            # Calculate max width
            max_length = len(str(col))
            for cell in worksheet[get_column_letter(idx)]:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Set width with some padding, max 60 characters
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        # Add autofilter
        if len(df) > 0:
            worksheet.auto_filter.ref = worksheet.dimensions
        
        # Freeze top row
        worksheet.freeze_panes = 'A2'
    
    def _create_errors_sheet(self, writer: pd.ExcelWriter, errors: List[str]):
        """Create the errors/skipped accounts sheet"""
        if not errors:
            # Create sheet with "No errors" message
            df = pd.DataFrame([{'Status': 'All accounts and regions scanned successfully!'}])
        else:
            df = pd.DataFrame({'Error/Skipped': errors})
        
        df.to_excel(writer, index=False, sheet_name='Errors-Skipped')
        
        # Format the sheet
        worksheet = writer.sheets['Errors-Skipped']
        
        # Style header
        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        cell = worksheet.cell(row=1, column=1)
        cell.fill = header_fill
        cell.font = header_font
        
        # Set column width
        worksheet.column_dimensions['A'].width = 100


def create_inventory_report(
    results: List[ScanResult],
    errors: List[str],
    output_dir: Path,
    accounts_scanned: int,
    regions_scanned: int
) -> List[Path]:
    """
    Convenience function to create the inventory Excel reports.
    
    Args:
        results: List of ScanResult objects
        errors: List of error strings
        output_dir: Directory to save the files
        accounts_scanned: Number of accounts scanned
        regions_scanned: Number of regions
        
    Returns:
        List of paths to created Excel files
    """
    exporter = ExcelExporter(output_dir)
    return exporter.export(results, errors, accounts_scanned, regions_scanned)
